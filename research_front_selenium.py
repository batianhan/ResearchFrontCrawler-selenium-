# ========#
#  设置  #
# ========#
# -*- coding:utf-8 -*-
import os, shutil
import time
import datetime
from urllib.parse import quote
from openpyxl import load_workbook
from selenium import webdriver


url = 'https://esi.clarivate.com/'
root, FILENAME=os.path.split(os.path.abspath(__file__))

# 用于暂存文件
temp_files = root + "\\temp_files"
if not os.path.exists(temp_files):
    os.mkdir(temp_files)

Fields = ['Agricultural Sciences', 'Biology & Biochemistry', 'Chemistry', 'Clinical Medicine', 'Computer Science', 'Economics & Business', 'Engineering', 'Environment/Ecology', 'Geosciences', 'Immunology', 'Materials Science', 'Mathematics', 'Microbiology', 'Molecular Biology & Genetics', 'Multidisciplinary', 'Neuroscience & Behavior', 'Pharmacology & Toxicology', 'Physics', 'Plant & Animal Science', 'Psychiatry/Psychology', 'Social Sciences, General', 'Space Science']
Fieldc = ['农业科学', '生物学与生物化学', '化学', '临床医学', '计算机科学', '经济学与商学', '工程学', '环境科学与生态学', '地球科学', '免疫学', '材料科学', '数学', '微生物学', '分子生物学与遗传学', '综合交叉学科', '神经学与行为学', '药理学与毒理学', '物理学', '动植物科学', '精神病学与心理学', '社会科学总论', '空间科学']
# test = 'EFFECTIVE GREEN CORROSION INHIBITOR;GREEN ENVIRONMENTAL CORROSION INHIBITORS;GREEN CORROSION INHIBITORS;GREEN CORROSION INHIBITOR;GREEN CORROSION INHIBITION'
month_short = {'Jan':1,'Feb':2,'Mar':3,'Apr':4,'May':5,'Jun':6,'Jul':7,'Aug':8,'Sep':9,'Oct':10,'Nov':11,'Dec':12}
years = time.localtime(time.time()).tm_year


# ========#
#  函数   #
# ========#
def chromeInit(savePath=os.getcwd(), flag=False):
    options=webdriver.ChromeOptions()
    options._binary_location= root + "/Application/chrome.exe"
    options.add_argument("--disable-gpu")
    options.add_argument('--allow-running-insecure-content')
    options.add_argument('--disable-extensions')

    if flag:
        options.add_argument('--headless')

    if savePath != '':
        prefs = {
            "profile.default_content_settings.popups": 0,
            "download.default_directory": savePath}
        options.add_experimental_option("prefs", prefs)
    chrome=webdriver.Chrome(options=options, executable_path=root + "/Application/chromedriver.exe")
    chrome.implicitly_wait(10)
    log_console("Chrome init complete...")
    return chrome


# 获取最新文件（时间排序取最后）
def sort_file(path):
    """排序文件"""
    dir_lists = os.listdir(path)
    dir_lists.sort(key=lambda fn: os.path.getmtime(path + '\\' + fn))
    return (dir_lists[-1])

# 移动文件，顺便改名
def movefile(src_file, dst_file):
    if not os.path.isfile(src_file):
        log_console("{} not exist!".format(src_file))
    else:
        dst_path,dst_fname=os.path.split(dst_file)     #分离文件名和路径
        if not os.path.exists(dst_path):   os.makedirs(dst_path)  #创建路径

        for fname in os.listdir(dst_path):  # 同名删除
            if fname == dst_fname: os.remove(dst_path + '\\' + fname)

        shutil.move(src_file, dst_file)         #移动文件
        # log_console("move {0} -> {1}".format(src_file, dst_file))

# 删除文件夹中的所有文件
def clean_fold(path):
    for i in os.listdir(path):              # os.listdir(path_data)#返回一个列表，里面是当前目录下面的所有东西的相对路径
        file = path + "\\" + i              # 当前文件夹的下面的所有东西的绝对路径
        if os.path.isfile(file) == True:    # os.path.isfile判断是否为文件,如果是文件,就删除.如果是文件夹.递归给del_file.
            os.remove(file)
        else:
            clean_fold(file)

# 用js点击
def js_click(driver, element):
    driver.execute_script("arguments[0].click();", element)

def IE(name, total):#IndicatorsExport
    return '''https://esi.clarivate.com/IndicatorsExport.action?exportFile&_dc=1368621151464&groupBy=ResearchFronts&start=0&limit={0}&filterBy=ResearchFields&filterValue={1}&show=Top&sort=%5B%7B%22property%22:%22highPapers%22,%22direction%22:%22DESC%22%7D%5D&colFilterVal=&exportType=indicators&colNames=RowSeq,,Research%20Fronts,Top%20Papers,Mean%20Year&fileType=Excel&f=IndicatorsExport.xls'''.format(total, quote(name.upper()))
def DE(name, limit_top_page):#DocumentsExport
    return '''https://esi.clarivate.com/DocumentsExport.action?exportFile&_dc=1368621151464&groupBy=documentList&start=0&limit={0}&author=&institution=&territory=&journal=&researchFront={1}&year=&title=&researchField=&show=Top&sort=%5B%7B"property"%3A"citations"%2C"direction"%3A"DESC"%7D%5D&colFilterVal=&exportType=documents&colNames=Accession Number,DOI,PMID,Article Name,Authors,Source,Research Field,Times Cited,Countries,Addresses,Institutions,Publication Date&fileType=Excel&f=DocumentsExport.xls'''.format(int(limit_top_page), quote(name.upper()))
def Download(chrome, urls, path):
    while True:
        try:
            log_console('访问此链接下载:{}'.format(urls))
            clean_fold(temp_files)
            chrome.get(urls)
        except Exception:
            log_console("网络错误（下载失败）\n")
            time.sleep(60)
            continue

        try:
            # 等待下载完成 确保中间文件(.tmp .crdownload)完全转好
            time.sleep(0.05)    # 测试出的合适时间
            # log_console('测试开始')
            while (len(os.listdir(temp_files)) == 0
                   or os.listdir(temp_files)[0].split('.')[-1] == 'tmp'
                   or os.listdir(temp_files)[0].split('.')[-1] == 'crdownload'): time.sleep(0.01)
            # log_console('测试结束')
            movefile(temp_files + '\\' + sort_file(temp_files), path)
            break
        except:
            log_console('文件移动转换错误\n')
            continue

# 打印信息时加上时间
def log_console(str):
    # print('{} {}'.format(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S.%f'), str))    # 含微秒的日期时间
    print('{} {}'.format(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'), str))
#=========#
# 主函数  #
#=========#
if __name__ == "__main__":
    log_console('已启动自动关机，将在21:55关机。')
    chrome = chromeInit(savePath=temp_files)
    log_console('正在加载...')
    while True:
        try:
            chrome.get(url)
            break
        except:
            log_console("{} 网络错误（esi加载失败）\n")
            time.sleep(60)

    update_str = chrome.find_element_by_css_selector("#updateDateDatasetESI").text
    t = update_str.split('updated ')[1]
    t = t.split(' ')[0]
    month = month_short[t]
    log_console('当前更新至{}月'.format(month))

    process_research_fronts = 0   #总进度
    for index_field in range(len(Fields)):
        try:
            f = open('./log/{}.{}/{}.log'.format(years, month, Fieldc[index_field]), 'r')
        except:
            process = 0
            if not os.path.exists('./{1}.{2}/{0}/'.format(Fieldc[index_field], years, month)):
                os.makedirs('./{1}.{2}/{0}/'.format(Fieldc[index_field], years, month))

            # ========== ========== ========== ==========
            # 获取学科条目总数（下载 Indicators 的时候链接里的limit需设置为总数）
            subject_total = 0
            # 选择Research Fronts并收集科目元素列表
            chrome.find_element_by_css_selector(".select2-choice").click()  # 点击弹出下拉菜单
            elements = chrome.find_elements_by_css_selector(".select2-result-label")  # 下拉菜单中的元素
            for element in elements:
                if element.text == "Research Fronts":
                    element.click()  # 选择 Research Fronts

            js_click(chrome, chrome.find_element_by_css_selector(".add-filters"))  # 点击添加过滤器
            chrome.find_element_by_css_selector(
                ".popup-wrapper>ul>li:nth-child(2)").click()  # 点击 Research Field
            subjectLabels = chrome.find_elements_by_css_selector(".checkbox-columns>div>label")  # 获取学科列表
            for subjectLabel in subjectLabels:
                if subjectLabel.text == Fields[index_field]:
                    subjectLabel.click()  # 选择学科
                    break
            time.sleep(1)  # 这个时间影响不大，每个学科只执行一次
            subject_total = int(chrome.find_element_by_css_selector("#grid>div:nth-child(1)>div>div>div:nth-child(1)>div>span")
                                .get_attribute('innerText')
                                .split(' ')[1])  # 获取该学科条目总数
            # ========== ========== ========== ==========

            Download(chrome, IE(Fields[index_field], subject_total), '{1}.{2}/{0}/{0}TOP{1}.{2}.xlsx'.format(Fieldc[index_field], years, month))
            log_console('{}Indicators获取完成'.format(Fieldc[index_field]))
        else:
            process = int(f.read())
            log_console('{}已收集至{}'.format(Fieldc[index_field], process))
            process_research_fronts += process
            f.close()

        ws = load_workbook('./{1}.{2}/{0}/{0}TOP{1}.{2}.xlsx'.format(Fieldc[index_field], years, month)).active
        total = ws.max_row - 7
        if process == total:
            log_console('{}已收集完成'.format(Fieldc[index_field]))
            continue
        row_range = ws[7+process:len(ws['A'])-1]

        for item in row_range:
            t = 0
            # print('下载{}'.format(item[1].value))
            Download(chrome, DE(item[1].value, item[2].value), './{}.{}/{}/{}-{}.xlsx'.format(years, month, Fieldc[index_field], process + 1, int(item[0].value)))
            try:
                data = load_workbook(
                    './{}.{}/{}/{}-{}.xlsx'.format(years, month, Fieldc[index_field], process + 1, int(item[0].value))).active
                keyword = data["A5"].value[43:-19]
            except:
                keyword = 'error'
            while keyword != item[1].value:
                log_console('数据有误，正在重试...')
                Download(chrome, DE(item[1].value, item[2].value), './{}.{}/{}/{}-{}.xlsx'.format(years, month, Fieldc[index_field], process + 1, int(item[0].value)))
                try:
                    data = load_workbook('./{}.{}/{}/{}-{}.xlsx'.format(years, month, Fieldc[index_field], process + 1, int(item[0].value))).active
                    keyword = data["A5"].value[43:-19]
                except:
                    continue
                t += 1
                if t >= 3:
                    log_console('{}下载失败'.format(item[1].value))
                    try:
                        f = open('./{}.{}/{}/failed.txt'.format(years, month, Fieldc[index_field]), 'a')
                        f.write('{}-{}\t{}\n'.format(process + 1, int(item[0].value), item[1].value))
                        f.close()
                    except:
                        f = open('./{}.{}/{}/failed.txt'.format(years, month, Fieldc[index_field]), 'w')
                        f.write('{}-{}\t{}\n'.format(process + 1, int(item[0].value), item[1].value))
                        f.close()
            if not os.path.exists('./log/{}.{}'.format(years, month)):
                os.makedirs('./log/{}.{}'.format(years, month))
            f = open('./log/{}.{}/{}.log'.format(years, month, Fieldc[index_field]), 'w')
            # keyword = data.iloc[3][0][43:-19]
            process += 1
            process_research_fronts += 1
            f.write(str(process))
            f.close()
            hour = time.localtime(time.time()).tm_hour
            minute = time.localtime(time.time()).tm_min
            if hour >= 21 and minute >= 55:
                log_console("关机...")
                os.system('shutdown -s')
            log_console('[{}/21]{}进度:{}/{} {:.2f}%\n'.format(index_field, Fieldc[index_field], process, total, process/total*100, process_research_fronts))
        log_console('{}收集完成\n'.format(Fieldc[index_field]))
    log_console('所有科目收集完成。')